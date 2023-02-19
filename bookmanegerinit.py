# -*- coding: utf-8 -*-
import requests
import json
import openpyxl
import glob
import bctest

url = 'https://www.googleapis.com/books/v1/volumes?q=isbn:'
wb = openpyxl.load_workbook("booklist.xlsx")
ws = wb.worksheets[0]
files = glob.glob("./bookpic/*.jpg")


def isbntoinfo(isbn):
    req_url = url + isbn
    response = requests.get(req_url).json()
    # totalitems = response['totalItems']  # 件数
    items_list = response['items']  # items リストデータ
    items = items_list[0]  # items
    info = items['volumeInfo']
    title = info['title']
    author = info['authors'][0]
    try:
        publisher = info['publisher']
    except Exception:
        publisher = "None"
    try:
        publisheddate = info['publishedDate']
    except Exception:
        publisheddate = "None"
    try:
        pages = info['pageCount']
    except Exception:
        pages = "None"
    try:
        language = info['language']
    except Exception:
        language = "None"
    # print(title)
    # print(author)
    # print(publisher)
    # print(language)
    return title, author, publisher, language, publisheddate, pages


def main():
    tmp = 2
    # isbn_input = input("input ISBN >>>")
    for path in files:
        try:
            isbn = bctest.bcdecoder(path)
            title, author, publisher, language, publisheddate, pages = isbntoinfo(
                isbn)
            ws.cell(row=tmp, column=2, value=isbn)
            ws.cell(row=tmp, column=3, value=title)
            ws.cell(row=tmp, column=4, value=author)
            ws.cell(row=tmp, column=5, value=publisher)
            ws.cell(row=tmp, column=6, value=language)
            ws.cell(row=tmp, column=7, value=publisheddate)
            ws.cell(row=tmp, column=8, value=pages)
            tmp += 1
        except Exception:
            pass
    wb.save("booklist.xlsx")
    wb.close()


if __name__ == "__main__":
    main()
